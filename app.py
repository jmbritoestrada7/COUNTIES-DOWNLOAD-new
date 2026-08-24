from __future__ import annotations

import json
import gzip
import os
import re
import uuid
import threading
import webbrowser
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
import boto3
from botocore.config import Config
from botocore.exceptions import BotoCoreError, ClientError
from openpyxl import load_workbook
from flask import Flask, jsonify, redirect, render_template, request, url_for
from flask_socketio import SocketIO, emit, join_room
from werkzeug.utils import secure_filename

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("DATA_DIR", str(BASE_DIR / "data")))
PROJECT_DIR = DATA_DIR / "projects"
UPLOAD_DIR = DATA_DIR / "uploads"
PROPERTY_DIR = DATA_DIR / "property_points"
PROPERTY_RECORD_DIR = DATA_DIR / "property_records"
MARKET_RECORD_DIR = DATA_DIR / "market_records"
MARKETING_ACTIVITY_DIR = DATA_DIR / "marketing_activity"
PROJECT_DIR.mkdir(parents=True, exist_ok=True)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
PROPERTY_DIR.mkdir(parents=True, exist_ok=True)
PROPERTY_RECORD_DIR.mkdir(parents=True, exist_ok=True)
MARKET_RECORD_DIR.mkdir(parents=True, exist_ok=True)
MARKETING_ACTIVITY_DIR.mkdir(parents=True, exist_ok=True)

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "local-development-secret")
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024
socketio = SocketIO(app, cors_allowed_origins="*", async_mode="threading")

STATE_FIPS = {
    "AL":"01","AK":"02","AZ":"04","AR":"05","CA":"06","CO":"08","CT":"09","DE":"10","DC":"11","FL":"12",
    "GA":"13","HI":"15","ID":"16","IL":"17","IN":"18","IA":"19","KS":"20","KY":"21","LA":"22","ME":"23",
    "MD":"24","MA":"25","MI":"26","MN":"27","MS":"28","MO":"29","MT":"30","NE":"31","NV":"32","NH":"33",
    "NJ":"34","NM":"35","NY":"36","NC":"37","ND":"38","OH":"39","OK":"40","OR":"41","PA":"42","RI":"44",
    "SC":"45","SD":"46","TN":"47","TX":"48","UT":"49","VT":"50","VA":"51","WA":"53","WV":"54","WI":"55","WY":"56"
}
FIPS_STATES = {v: k for k, v in STATE_FIPS.items()}

STATE_NAMES = {
    "ALABAMA":"AL","ALASKA":"AK","ARIZONA":"AZ","ARKANSAS":"AR","CALIFORNIA":"CA","COLORADO":"CO","CONNECTICUT":"CT",
    "DELAWARE":"DE","DISTRICT OF COLUMBIA":"DC","FLORIDA":"FL","GEORGIA":"GA","HAWAII":"HI","IDAHO":"ID","ILLINOIS":"IL",
    "INDIANA":"IN","IOWA":"IA","KANSAS":"KS","KENTUCKY":"KY","LOUISIANA":"LA","MAINE":"ME","MARYLAND":"MD",
    "MASSACHUSETTS":"MA","MICHIGAN":"MI","MINNESOTA":"MN","MISSISSIPPI":"MS","MISSOURI":"MO","MONTANA":"MT","NEBRASKA":"NE",
    "NEVADA":"NV","NEW HAMPSHIRE":"NH","NEW JERSEY":"NJ","NEW MEXICO":"NM","NEW YORK":"NY","NORTH CAROLINA":"NC",
    "NORTH DAKOTA":"ND","OHIO":"OH","OKLAHOMA":"OK","OREGON":"OR","PENNSYLVANIA":"PA","RHODE ISLAND":"RI",
    "SOUTH CAROLINA":"SC","SOUTH DAKOTA":"SD","TENNESSEE":"TN","TEXAS":"TX","UTAH":"UT","VERMONT":"VT","VIRGINIA":"VA",
    "WASHINGTON":"WA","WEST VIRGINIA":"WV","WISCONSIN":"WI","WYOMING":"WY"
}


PROJECT_SCHEMA_VERSION = 11


def migrate_project(project: dict) -> tuple[dict, bool]:
    """Upgrade older stored projects without changing their permanent links."""
    changed = False
    if not isinstance(project, dict):
        raise ValueError("Invalid project data")

    version = int(project.get("schema_version") or 1)
    project.setdefault("counties", [])
    project.setdefault("drawings", {"type": "FeatureCollection", "features": []})
    project.setdefault("property_analytics", {"loaded": False, "source_files": [], "total_properties": 0, "states": []})
    project.setdefault("market_analytics", {"loaded": False, "active_files": [], "sold_files": [], "active_count": 0, "sold_count": 0, "states": []})
    project.setdefault("marketing_activity", {"loaded": False, "source_files": [], "states": [], "events": 0})

    drawings = project.get("drawings")
    if not isinstance(drawings, dict) or drawings.get("type") != "FeatureCollection":
        project["drawings"] = {"type": "FeatureCollection", "features": []}
        drawings = project["drawings"]
        changed = True

    for index, feature in enumerate(drawings.get("features") or [], start=1):
        if not isinstance(feature, dict):
            continue
        props = feature.setdefault("properties", {})
        if not props.get("id"):
            props["id"] = uuid.uuid4().hex
            changed = True
        if not str(props.get("name") or "").strip():
            props["name"] = f"Unnamed Area {index}"
            changed = True
        if not props.get("shapeType"):
            geom_type = str((feature.get("geometry") or {}).get("type") or "")
            props["shapeType"] = "Circle" if geom_type == "Point" and props.get("radius") else "Polygon"
            changed = True
        if not props.get("color"):
            props["color"] = "#7c3aed"
            changed = True
        if "visible" not in props:
            props["visible"] = True
            changed = True

    default_settings = {
        "state_filter": "",
        "str_min": None,
        "str_max": None,
        "search_filter": "",
        "color_metric": "str_value",
        "property_color_mode": "automatic",
        "property_thresholds": [25, 100, 250, 500],
        "property_point_style": {
            "size": 4,
            "color": "#22c55e",
            "opacity": 0.75,
            "outline_color": "#0f172a",
            "outline_width": 1,
        },
        "layers": {
            "counties": True,
            "county_labels": True,
            "str_colors": True,
            "drawings": True,
            "drawing_labels": True,
            "property_points": True,
        },
        "layer_order": ["counties", "drawings", "drawing_labels", "county_labels", "property_points"],
        "county_card_fields": [
            "average_str","property_count","unique_owner_count","portfolio_owner_count","portfolio_property_count",
            "total_acreage","average_acreage","market_active_count","market_sold_count","market_str",
            "market_avg_active_price","market_avg_sold_price","str_by_acreage",
            "mailer_sent","rvm","ai_texting","cold_calling","neutral_postcard"
        ],
    }
    settings = project.setdefault("view_settings", {})
    for key, value in default_settings.items():
        if key not in settings:
            settings[key] = value.copy() if isinstance(value, dict) else value
            changed = True
    layers = settings.setdefault("layers", {})
    for key, value in default_settings["layers"].items():
        if key not in layers:
            layers[key] = value
            changed = True

    if version != PROJECT_SCHEMA_VERSION:
        project["schema_version"] = PROJECT_SCHEMA_VERSION
        changed = True
    return project, changed


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def project_path(project_id: str) -> Path:
    if not re.fullmatch(r"[a-f0-9]{12}", project_id):
        raise ValueError("Invalid project id")
    return PROJECT_DIR / f"{project_id}.json"


def validate_project_id(project_id: str) -> str:
    if not re.fullmatch(r"[a-f0-9]{12}", str(project_id or "")):
        raise ValueError("Invalid project id")
    return project_id


class ProjectStorage:
    """Persist project JSON locally and, when configured, in Cloudflare R2.

    Render's normal filesystem is ephemeral. R2 is the source of truth in
    production; the local copy is only a fast cache and a development fallback.
    """

    def __init__(self) -> None:
        self.account_id = os.environ.get("R2_ACCOUNT_ID", "").strip()
        self.access_key = os.environ.get("R2_ACCESS_KEY_ID", "").strip()
        self.secret_key = os.environ.get("R2_SECRET_ACCESS_KEY", "").strip()
        self.bucket = os.environ.get("R2_BUCKET_NAME", "").strip()
        self.prefix = os.environ.get("R2_PROJECT_PREFIX", "projects").strip().strip("/") or "projects"
        self.endpoint = os.environ.get("R2_ENDPOINT_URL", "").strip()
        self.client = None

        configured = all([self.account_id, self.access_key, self.secret_key, self.bucket])
        if configured:
            endpoint_url = self.endpoint or f"https://{self.account_id}.r2.cloudflarestorage.com"
            self.client = boto3.client(
                service_name="s3",
                endpoint_url=endpoint_url,
                aws_access_key_id=self.access_key,
                aws_secret_access_key=self.secret_key,
                region_name="auto",
                config=Config(signature_version="s3v4", retries={"max_attempts": 4, "mode": "standard"}),
            )

    @property
    def mode(self) -> str:
        return "r2" if self.client else "local"

    def key(self, project_id: str) -> str:
        validate_project_id(project_id)
        return f"{self.prefix}/{project_id}.json"

    def save(self, project: dict) -> None:
        project_id = validate_project_id(str(project.get("id", "")))
        payload = json.dumps(project, ensure_ascii=False, indent=2).encode("utf-8")

        # Always keep a local cache so local development works without R2.
        project_path(project_id).write_bytes(payload)

        if self.client:
            try:
                self.client.put_object(
                    Bucket=self.bucket,
                    Key=self.key(project_id),
                    Body=payload,
                    ContentType="application/json; charset=utf-8",
                    CacheControl="no-store",
                )
            except (BotoCoreError, ClientError) as exc:
                raise RuntimeError(f"Could not save the project to Cloudflare R2: {exc}") from exc

    def load(self, project_id: str) -> dict:
        validate_project_id(project_id)
        if self.client:
            try:
                response = self.client.get_object(Bucket=self.bucket, Key=self.key(project_id))
                project = json.loads(response["Body"].read().decode("utf-8"))
                # Refresh the local cache after a successful remote read.
                project_path(project_id).write_text(
                    json.dumps(project, ensure_ascii=False, indent=2), encoding="utf-8"
                )
                return project
            except self.client.exceptions.NoSuchKey:
                pass
            except ClientError as exc:
                code = str(exc.response.get("Error", {}).get("Code", ""))
                if code not in {"NoSuchKey", "404", "NotFound"}:
                    raise RuntimeError(f"Could not read the project from Cloudflare R2: {exc}") from exc
            except BotoCoreError as exc:
                raise RuntimeError(f"Could not connect to Cloudflare R2: {exc}") from exc

        path = project_path(project_id)
        if not path.exists():
            raise FileNotFoundError(project_id)
        return json.loads(path.read_text(encoding="utf-8"))

    def delete(self, project_id: str) -> None:
        validate_project_id(project_id)
        path = project_path(project_id)
        if path.exists():
            path.unlink()
        if self.client:
            try:
                self.client.delete_object(Bucket=self.bucket, Key=self.key(project_id))
            except (BotoCoreError, ClientError) as exc:
                raise RuntimeError(f"Could not delete the project from Cloudflare R2: {exc}") from exc

    def property_key(self, project_id: str) -> str:
        validate_project_id(project_id)
        return f"property_points/{project_id}.json.gz"

    def save_property_points(self, project_id: str, points: list[dict]) -> None:
        validate_project_id(project_id)
        payload = gzip.compress(json.dumps(points, ensure_ascii=False, separators=(",", ":")).encode("utf-8"), compresslevel=6)
        local = PROPERTY_DIR / f"{project_id}.json.gz"
        local.write_bytes(payload)
        if self.client:
            self.client.put_object(Bucket=self.bucket, Key=self.property_key(project_id), Body=payload, ContentType="application/json", ContentEncoding="gzip", CacheControl="no-store")

    def load_property_points(self, project_id: str) -> list[dict]:
        validate_project_id(project_id)
        payload = None
        if self.client:
            try:
                response = self.client.get_object(Bucket=self.bucket, Key=self.property_key(project_id))
                payload = response["Body"].read()
                (PROPERTY_DIR / f"{project_id}.json.gz").write_bytes(payload)
            except ClientError as exc:
                code = str(exc.response.get("Error", {}).get("Code", ""))
                if code not in {"NoSuchKey", "404", "NotFound"}:
                    raise
        if payload is None:
            local = PROPERTY_DIR / f"{project_id}.json.gz"
            if not local.exists():
                return []
            payload = local.read_bytes()
        return json.loads(gzip.decompress(payload).decode("utf-8"))

    def property_records_key(self, project_id: str) -> str:
        validate_project_id(project_id)
        return f"property_records/{project_id}.json.gz"

    def save_property_records(self, project_id: str, records: list[dict]) -> None:
        validate_project_id(project_id)
        payload = gzip.compress(json.dumps(records, ensure_ascii=False, separators=(",", ":")).encode("utf-8"), compresslevel=6)
        local = PROPERTY_RECORD_DIR / f"{project_id}.json.gz"
        local.write_bytes(payload)
        if self.client:
            self.client.put_object(Bucket=self.bucket, Key=self.property_records_key(project_id), Body=payload, ContentType="application/json", ContentEncoding="gzip", CacheControl="no-store")

    def load_property_records(self, project_id: str) -> list[dict]:
        validate_project_id(project_id)
        payload = None
        if self.client:
            try:
                response = self.client.get_object(Bucket=self.bucket, Key=self.property_records_key(project_id))
                payload = response["Body"].read()
                (PROPERTY_RECORD_DIR / f"{project_id}.json.gz").write_bytes(payload)
            except ClientError as exc:
                code = str(exc.response.get("Error", {}).get("Code", ""))
                if code not in {"NoSuchKey", "404", "NotFound"}:
                    raise
        if payload is None:
            local = PROPERTY_RECORD_DIR / f"{project_id}.json.gz"
            if not local.exists():
                return []
            payload = local.read_bytes()
        return json.loads(gzip.decompress(payload).decode("utf-8"))

    def market_records_key(self, project_id: str) -> str:
        validate_project_id(project_id)
        return f"market_records/{project_id}.json.gz"

    def save_market_records(self, project_id: str, records: list[dict]) -> None:
        validate_project_id(project_id)
        payload = gzip.compress(json.dumps(records, ensure_ascii=False, separators=(",", ":")).encode("utf-8"), compresslevel=6)
        local = MARKET_RECORD_DIR / f"{project_id}.json.gz"
        local.write_bytes(payload)
        if self.client:
            self.client.put_object(Bucket=self.bucket, Key=self.market_records_key(project_id), Body=payload, ContentType="application/json", ContentEncoding="gzip", CacheControl="no-store")

    def load_market_records(self, project_id: str) -> list[dict]:
        validate_project_id(project_id)
        payload = None
        if self.client:
            try:
                response = self.client.get_object(Bucket=self.bucket, Key=self.market_records_key(project_id))
                payload = response["Body"].read()
                (MARKET_RECORD_DIR / f"{project_id}.json.gz").write_bytes(payload)
            except ClientError as exc:
                code = str(exc.response.get("Error", {}).get("Code", ""))
                if code not in {"NoSuchKey", "404", "NotFound"}:
                    raise
        if payload is None:
            local = MARKET_RECORD_DIR / f"{project_id}.json.gz"
            if not local.exists():
                return []
            payload = local.read_bytes()
        return json.loads(gzip.decompress(payload).decode("utf-8"))

    def marketing_activity_key(self, project_id: str) -> str:
        validate_project_id(project_id)
        return f"marketing_activity/{project_id}.json.gz"

    def save_marketing_activity(self, project_id: str, records: list[dict]) -> None:
        validate_project_id(project_id)
        payload = gzip.compress(json.dumps(records, ensure_ascii=False, separators=(",", ":")).encode("utf-8"), compresslevel=6)
        local = MARKETING_ACTIVITY_DIR / f"{project_id}.json.gz"
        local.write_bytes(payload)
        if self.client:
            self.client.put_object(Bucket=self.bucket, Key=self.marketing_activity_key(project_id), Body=payload, ContentType="application/json", ContentEncoding="gzip", CacheControl="no-store")

    def load_marketing_activity(self, project_id: str) -> list[dict]:
        validate_project_id(project_id)
        payload = None
        if self.client:
            try:
                response = self.client.get_object(Bucket=self.bucket, Key=self.marketing_activity_key(project_id))
                payload = response["Body"].read()
                (MARKETING_ACTIVITY_DIR / f"{project_id}.json.gz").write_bytes(payload)
            except ClientError as exc:
                code = str(exc.response.get("Error", {}).get("Code", ""))
                if code not in {"NoSuchKey", "404", "NotFound"}:
                    raise
        if payload is None:
            local = MARKETING_ACTIVITY_DIR / f"{project_id}.json.gz"
            if not local.exists():
                return []
            payload = local.read_bytes()
        return json.loads(gzip.decompress(payload).decode("utf-8"))

    def list_projects(self) -> list[dict]:
        projects_by_id: dict[str, dict] = {}

        if self.client:
            token = None
            while True:
                args = {"Bucket": self.bucket, "Prefix": f"{self.prefix}/", "MaxKeys": 1000}
                if token:
                    args["ContinuationToken"] = token
                try:
                    result = self.client.list_objects_v2(**args)
                except (BotoCoreError, ClientError) as exc:
                    raise RuntimeError(f"Could not list projects from Cloudflare R2: {exc}") from exc
                for obj in result.get("Contents", []):
                    key = str(obj.get("Key", ""))
                    match = re.fullmatch(rf"{re.escape(self.prefix)}/([a-f0-9]{{12}})\.json", key)
                    if not match:
                        continue
                    project_id = match.group(1)
                    try:
                        project = self.load(project_id)
                        projects_by_id[project_id] = project
                    except Exception:
                        continue
                if not result.get("IsTruncated"):
                    break
                token = result.get("NextContinuationToken")

        # Include local-only projects when running without R2 or during migration.
        for path in PROJECT_DIR.glob("*.json"):
            try:
                project = json.loads(path.read_text(encoding="utf-8"))
                projects_by_id.setdefault(project["id"], project)
            except Exception:
                continue

        return sorted(projects_by_id.values(), key=lambda p: p.get("updated_at", ""), reverse=True)


storage = ProjectStorage()


def load_project(project_id: str) -> dict:
    project, changed = migrate_project(storage.load(project_id))
    if changed:
        project["updated_at"] = now_iso()
        storage.save(project)
    return project


def save_project(project: dict) -> None:
    project, _ = migrate_project(project)
    project["updated_at"] = now_iso()
    storage.save(project)


def normalize_state(value) -> tuple[str, str]:
    raw = str(value or "").strip().upper()
    abbr = raw if raw in STATE_FIPS else STATE_NAMES.get(raw, "")
    return abbr, STATE_FIPS.get(abbr, "")


def normalize_county(value) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\s+(county|parish|borough|census area|municipality|city and borough)$", "", text, flags=re.I)
    return re.sub(r"\s+", " ", text).strip()


def find_column(df: pd.DataFrame, aliases: list[str]):
    normalized = {re.sub(r"[^a-z0-9]", "", str(c).lower()): c for c in df.columns}
    for alias in aliases:
        key = re.sub(r"[^a-z0-9]", "", alias.lower())
        if key in normalized:
            return normalized[key]
    return None



def parse_str(value, percent_formatted: bool = False) -> tuple[str, float | None]:
    """Return a display percentage and numeric percent value.

    Excel percentage-formatted cells store 134% as 1.34. When the cell format
    contains %, multiply by 100. Plain numeric values such as 134 remain 134.
    Text values ending in % are already expressed as percentages.
    """
    if value is None or value == "":
        return "", None
    raw = str(value).strip().replace(",", ".")
    if not raw:
        return "", None
    has_symbol = "%" in raw
    cleaned = re.sub(r"[^0-9.\-]", "", raw)
    try:
        number = float(cleaned)
    except (TypeError, ValueError):
        return raw, None
    if percent_formatted and not has_symbol:
        number *= 100
    elif not has_symbol and 0 <= number < 1:
        number *= 100
    formatted = f"{number:.2f}".rstrip("0").rstrip(".")
    return f"{formatted}%", number


STR_COLUMN_ALIASES = {
    "str_2_5": ["STR 2-5", "SRT 2-5", "STR_2_5", "STR 2 TO 5", "2-5"],
    "str_5_10": ["STR 5-10", "SRT 5-10", "STR_5_10", "STR 5 TO 10", "5-10"],
    "str_10_20": ["STR 10-20", "SRT 10-20", "STR_10_20", "STR 10 TO 20", "10-20"],
    "str_20_60": ["STR 20-60", "SRT 20-60", "STR_20_60", "STR 20 TO 60", "20-60"],
    "str_60_100": ["STR 60-100", "SRT 60-100", "STR_60_100", "STR 60 TO 100", "60-100"],
    "str_100_plus": ["STR 100+", "SRT 100+", "SRT 100- +", "STR 100- +", "STR_100_PLUS", "STR 100 PLUS", "100+"],
}


def county_identity(row: dict) -> tuple[str, str]:
    return str(row.get("state_fips", "")), str(row.get("county_key", ""))


def merge_existing_notes(new_counties: list[dict], old_counties: list[dict]) -> list[dict]:
    old_data = {county_identity(c): c for c in old_counties}
    for county in new_counties:
        previous = old_data.get(county_identity(county), {})
        for field in ("notes", "priority", "assigned_to", "next_review"):
            if not str(county.get(field, "")).strip() and previous.get(field):
                county[field] = previous[field]
        preserve_fields = ["property_count","unique_owner_count","portfolio_owner_count","portfolio_property_count","total_acreage","average_acreage"]
        for ch in ("mailer_sent","rvm","ai_texting","cold_calling","neutral_postcard"):
            preserve_fields += [f"marketing_{ch}_count", f"marketing_{ch}_dates"]
        for field in preserve_fields:
            if field not in county and field in previous:
                county[field] = previous[field]
    return new_counties


def read_counties_excel(path: Path) -> list[dict]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    sheet = workbook.active
    header_values = [cell.value for cell in sheet[1]]
    df_headers = pd.DataFrame(columns=[str(v or "").strip() for v in header_values])

    county_col = find_column(df_headers, ["county", "county name", "condado", "county_name"])
    state_col = find_column(df_headers, ["state", "state code", "estado", "st"])
    status_col = find_column(df_headers, ["status", "estado descarga", "download status", "descargado"])
    date_col = find_column(df_headers, ["date", "download date", "fecha", "fecha descarga"])
    notes_col = find_column(df_headers, ["notes", "nota", "notas", "comments", "comentarios"])
    priority_col = find_column(df_headers, ["priority", "prioridad"])
    assigned_col = find_column(df_headers, ["assigned to", "assigned", "asignado a", "responsable"])
    review_col = find_column(df_headers, ["next review", "review date", "proxima revision", "próxima revisión"])
    str_col = find_column(df_headers, ["str", "sell through rate", "sell-through rate", "sell through", "tasa de venta", "porcentaje de venta", "avg str", "average of str"])
    str_band_cols = {key: find_column(df_headers, aliases) for key, aliases in STR_COLUMN_ALIASES.items()}
    if not county_col or not state_col:
        raise ValueError("The Excel file must include COUNTY and STATE columns.")

    header_index = {str(v or "").strip(): i + 1 for i, v in enumerate(header_values)}

    def cell_value(row_number: int, column_name: str | None):
        if not column_name:
            return "", False
        cell = sheet.cell(row=row_number, column=header_index[column_name])
        return cell.value, "%" in str(cell.number_format or "")

    rows = []
    for row_number in range(2, sheet.max_row + 1):
        county_value, _ = cell_value(row_number, county_col)
        state_value, _ = cell_value(row_number, state_col)
        county = normalize_county(county_value)
        state, state_fips = normalize_state(state_value)
        if not county or not state_fips:
            continue

        avg_raw, avg_pct = cell_value(row_number, str_col)
        avg_display, avg_value = parse_str(avg_raw, avg_pct)
        band_data = {}
        for key, column_name in str_band_cols.items():
            raw, pct = cell_value(row_number, column_name)
            display, numeric = parse_str(raw, pct)
            band_data[key] = display
            band_data[f"{key}_value"] = numeric

        status_value, _ = cell_value(row_number, status_col)
        date_value, _ = cell_value(row_number, date_col)
        notes_value, _ = cell_value(row_number, notes_col)
        priority_value, _ = cell_value(row_number, priority_col)
        assigned_value, _ = cell_value(row_number, assigned_col)
        review_value, _ = cell_value(row_number, review_col)

        rows.append({
            "county": county,
            "county_key": county.casefold(),
            "state": state,
            "state_fips": state_fips,
            "status": str(status_value or "Downloaded").strip() or "Downloaded",
            "date": str(date_value or "").strip(),
            "notes": str(notes_value or "").strip(),
            "priority": str(priority_value or "").strip(),
            "assigned_to": str(assigned_value or "").strip(),
            "next_review": str(review_value or "").strip(),
            "str": avg_display,
            "str_value": avg_value,
            **band_data,
        })
    if not rows:
        raise ValueError("No valid county and state rows were found.")
    return rows


def read_counties_csv(path: Path) -> list[dict]:
    """Read a CSV containing at minimum County and State columns.

    STR columns are optional. Counties without STR are intentionally retained
    and displayed with a neutral gray color in the map.
    """
    last_error = None
    df = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            df = pd.read_csv(path, dtype=object, encoding=encoding)
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    if df is None:
        raise ValueError(f"Could not read the CSV file: {last_error}")

    county_col = find_column(df, ["county", "county name", "condado", "county_name"])
    state_col = find_column(df, ["state", "state code", "estado", "st"])
    status_col = find_column(df, ["status", "download status", "descargado"])
    date_col = find_column(df, ["date", "download date", "fecha", "fecha descarga"])
    notes_col = find_column(df, ["notes", "nota", "notas", "comments", "comentarios"])
    priority_col = find_column(df, ["priority", "prioridad"])
    assigned_col = find_column(df, ["assigned to", "assigned", "asignado a", "responsable"])
    review_col = find_column(df, ["next review", "review date", "proxima revision", "próxima revisión"])
    str_col = find_column(df, ["str", "sell through rate", "sell-through rate", "sell through", "avg str", "average of str"])
    str_band_cols = {key: find_column(df, aliases) for key, aliases in STR_COLUMN_ALIASES.items()}
    if not county_col or not state_col:
        raise ValueError("The CSV file must include COUNTY and STATE columns.")

    def value(row, col):
        if not col:
            return ""
        val = row.get(col, "")
        return "" if pd.isna(val) else val

    rows = []
    for _, row in df.iterrows():
        county = normalize_county(value(row, county_col))
        state, state_fips = normalize_state(value(row, state_col))
        if not county or not state_fips:
            continue
        avg_display, avg_value = parse_str(value(row, str_col), False)
        band_data = {}
        for key, col in str_band_cols.items():
            display, numeric = parse_str(value(row, col), False)
            band_data[key] = display
            band_data[f"{key}_value"] = numeric
        rows.append({
            "county": county, "county_key": county.casefold(), "state": state, "state_fips": state_fips,
            "status": str(value(row, status_col) or "Selected").strip() or "Selected",
            "date": str(value(row, date_col) or "").strip(),
            "notes": str(value(row, notes_col) or "").strip(),
            "priority": str(value(row, priority_col) or "").strip(),
            "assigned_to": str(value(row, assigned_col) or "").strip(),
            "next_review": str(value(row, review_col) or "").strip(),
            "str": avg_display, "str_value": avg_value, **band_data,
        })
    if not rows:
        raise ValueError("No valid county and state rows were found.")
    return rows


def _read_tabular_file(path: Path) -> pd.DataFrame:
    ext = path.suffix.lower()
    if ext in {".xlsx", ".xlsm"}:
        return pd.read_excel(path, dtype=object)
    last = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return pd.read_csv(path, dtype=object, encoding=enc, low_memory=False)
        except UnicodeDecodeError as exc:
            last = exc
    raise ValueError(f"Could not read property file: {last}")


def _safe_float(value):
    if value is None or pd.isna(value):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except Exception:
        return None


def _norm_owner_piece(value: str) -> str:
    text = str(value or "").strip().casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _norm_ref(value: str) -> str:
    # REF is the authoritative property identifier. Preserve punctuation because
    # it may be meaningful; only normalize whitespace/case and Excel's numeric .0.
    text = str(value or "").strip().casefold()
    if re.fullmatch(r"[-+]?\d+\.0+", text):
        text = text.split(".", 1)[0]
    return text


def read_property_file(path: Path) -> tuple[list[dict], dict]:
    """Read an upload into canonical property records.

    REF is the immutable property key.  A future upload with the same REF updates
    the existing record instead of adding another property.
    """
    df = _read_tabular_file(path)
    state_col = find_column(df, ["state", "property state", "prop state", "estado", "st"])
    county_col = find_column(df, ["county", "county name", "property county", "condado"])
    lat_col = find_column(df, ["lat", "latitude", "property lat", "property latitude"])
    lon_col = find_column(df, ["long", "lon", "longitude", "property long", "property longitude"])
    acres_col = find_column(df, ["acreage", "acreage final", "acres", "cal acreage", "aggr acreage"])
    ref_col = find_column(df, ["ref", "property ref", "property_ref", "reference", "reference id"])
    apn_col = find_column(df, ["apn", "parcel id", "parcel_id", "parcel apn"])
    first_col = find_column(df, ["first_name", "first name", "owner first", "owner_1_first_n", "owner_1_first"])
    last_col = find_column(df, ["last_name", "last name", "owner last", "owner_1_last_n", "owner_1_last", "owner name"])
    address_col = find_column(df, ["mailing address", "mail_addr_n", "mail addr", "owner address", "mail_address"])
    city_col = find_column(df, ["mailing_city", "mail city", "mail_city_n", "owner city"])
    mail_state_col = find_column(df, ["mailing_state", "mail state", "mail_state_n", "owner state"])
    zip_col = find_column(df, ["zipcode", "zip", "mail_zip_5", "mail zip"])
    if not state_col or not county_col:
        raise ValueError("Property Analytics requires STATE and COUNTY columns. LAT/LONG are used for point display.")
    if not ref_col:
        raise ValueError("Property Analytics now requires a REF column. REF is used as the unique property ID so repeat uploads update instead of duplicate properties.")

    def val(row, col):
        if not col:
            return ""
        v = row.get(col, "")
        return "" if pd.isna(v) else str(v).strip()

    records_by_ref = {}
    skipped_no_ref = 0
    duplicate_refs_in_file = 0
    for _, row in df.iterrows():
        ref = val(row, ref_col)
        if not ref:
            skipped_no_ref += 1
            continue
        ref_key = _norm_ref(ref)
        if not ref_key:
            skipped_no_ref += 1
            continue
        state, sf = normalize_state(val(row, state_col))
        county = normalize_county(val(row, county_col))
        if not sf or not county:
            continue
        first, last = val(row, first_col), val(row, last_col)
        address, city, mst, zc = val(row, address_col), val(row, city_col), val(row, mail_state_col), val(row, zip_col)
        owner_parts = [_norm_owner_piece(x) for x in (first, last, address, city, mst, zc)]
        owner_key = "|".join(owner_parts)
        # Name + mailing address is the owner identity.  City/state/ZIP are retained
        # in the key as extra protection against common names and bad abbreviations.
        if not any(owner_parts[:2]) or not owner_parts[2]:
            owner_key = f"REFOWNER|{ref_key}"
        record = {
            "ref": ref,
            "ref_key": ref_key,
            "state": state,
            "state_fips": sf,
            "county": county,
            "county_key": county.casefold(),
            "lat": _safe_float(row.get(lat_col)) if lat_col else None,
            "lon": _safe_float(row.get(lon_col)) if lon_col else None,
            "acres": _safe_float(row.get(acres_col)) if acres_col else None,
            "apn": val(row, apn_col),
            "owner": (" ".join(x for x in [first, last] if x)).strip() or last,
            "owner_key": owner_key,
            "mailing_address": address,
            "mailing_city": city,
            "mailing_state": mst,
            "zipcode": zc,
        }
        if ref_key in records_by_ref:
            duplicate_refs_in_file += 1
        # Last occurrence in the upload wins, which is deterministic and prevents duplicates.
        records_by_ref[ref_key] = record

    records = list(records_by_ref.values())
    if not records:
        raise ValueError("No valid property rows with REF, STATE, and COUNTY were found.")
    meta = {
        "rows_uploaded": int(len(df)),
        "valid_refs": len(records),
        "duplicate_refs_in_file": duplicate_refs_in_file,
        "skipped_no_ref": skipped_no_ref,
        "states": sorted(set(r["state"] for r in records)),
    }
    return records, meta


def merge_property_records(existing: list[dict], incoming: list[dict]) -> tuple[list[dict], dict]:
    """UPSERT property records using REF as the one-and-only property key."""
    by_ref = {str(r.get("ref_key") or ""): dict(r) for r in existing if r.get("ref_key")}
    new_count = 0
    updated_count = 0
    unchanged_count = 0
    now = now_iso()
    for record in incoming:
        key = record["ref_key"]
        old = by_ref.get(key)
        if old is None:
            record = dict(record)
            record["first_seen_at"] = now
            record["last_updated_at"] = now
            by_ref[key] = record
            new_count += 1
        else:
            preserved_first_seen = old.get("first_seen_at") or now
            comparable_old = {k: v for k, v in old.items() if k not in {"first_seen_at", "last_updated_at"}}
            comparable_new = dict(record)
            record = dict(record)
            record["first_seen_at"] = preserved_first_seen
            record["last_updated_at"] = now
            by_ref[key] = record
            if comparable_old == comparable_new:
                unchanged_count += 1
            else:
                updated_count += 1
    merged = sorted(by_ref.values(), key=lambda r: (r.get("state", ""), r.get("county", ""), r.get("ref_key", "")))
    return merged, {"new": new_count, "updated": updated_count, "unchanged": unchanged_count}


def build_property_metrics(records: list[dict]) -> tuple[list[dict], list[dict], dict]:
    owner_counts = {}
    for r in records:
        owner_counts[r["owner_key"]] = owner_counts.get(r["owner_key"], 0) + 1
    grouped = {}
    for r in records:
        k = (r["state_fips"], r["county_key"])
        g = grouped.setdefault(k, {"state": r["state"], "state_fips": r["state_fips"], "county": r["county"], "county_key": r["county_key"], "rows": [], "owners": set(), "portfolio_owners": set()})
        g["rows"].append(r)
        g["owners"].add(r["owner_key"])
        if owner_counts[r["owner_key"]] > 1:
            g["portfolio_owners"].add(r["owner_key"])
    metrics = []
    for g in grouped.values():
        acreage = [r["acres"] for r in g["rows"] if r.get("acres") is not None]
        portfolio_props = sum(1 for r in g["rows"] if owner_counts[r["owner_key"]] > 1)
        metrics.append({
            "state": g["state"], "state_fips": g["state_fips"], "county": g["county"], "county_key": g["county_key"],
            "property_count": len(g["rows"]), "unique_owner_count": len(g["owners"]), "portfolio_owner_count": len(g["portfolio_owners"]),
            "portfolio_property_count": portfolio_props, "total_acreage": round(sum(acreage), 2), "average_acreage": round(sum(acreage) / len(acreage), 2) if acreage else None,
        })
    metrics.sort(key=lambda m: (m["state"], m["county"]))
    points = []
    for r in records:
        if r.get("lat") is None or r.get("lon") is None or not (-90 <= r["lat"] <= 90 and -180 <= r["lon"] <= 180):
            continue
        points.append({k: r.get(k) for k in ("ref", "state", "state_fips", "county", "county_key", "lat", "lon", "acres", "apn", "owner")})
    meta = {
        "total_properties": len(records),
        "unique_owners": len(owner_counts),
        "portfolio_owners": sum(1 for n in owner_counts.values() if n > 1),
        "point_count": len(points),
        "states": sorted(set(r["state"] for r in records)),
        "county_count": len(metrics),
    }
    return metrics, points, meta


def merge_property_metrics_into_counties(project: dict, metrics: list[dict]) -> list[dict]:
    counties = project.setdefault("counties", [])
    metric_fields = ("property_count", "unique_owner_count", "portfolio_owner_count", "portfolio_property_count", "total_acreage", "average_acreage")
    # Metrics are rebuilt from the complete REF master on every upload, so clear old
    # analytics first. STR/status/notes/drawings are untouched.
    for c in counties:
        for field in metric_fields:
            c.pop(field, None)
    by_id = {(c.get("state_fips"), c.get("county_key")): c for c in counties}
    for m in metrics:
        key = (m["state_fips"], m["county_key"])
        c = by_id.get(key)
        if c is None:
            c = {"county": m["county"], "county_key": m["county_key"], "state": m["state"], "state_fips": m["state_fips"], "status": "Property data", "date": "", "notes": "", "priority": "", "assigned_to": "", "next_review": "", "str": "", "str_value": None}
            for band in STR_COLUMN_ALIASES:
                c[band] = ""
                c[f"{band}_value"] = None
            counties.append(c)
            by_id[key] = c
        for field in metric_fields:
            c[field] = m.get(field)
    counties.sort(key=lambda c: (c.get("state", ""), c.get("county", "")))
    return counties



def _money_float(value):
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    text = re.sub(r"[^0-9.\-]", "", text.replace(",", ""))
    try:
        return float(text) if text not in {"", "-", "."} else None
    except Exception:
        return None


def _acres_float(value):
    if value is None or pd.isna(value):
        return None
    text = str(value).strip().lower().replace(",", "")
    m = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    return float(m.group(0)) if m else None


MARKETING_CHANNELS = {
    "mailer_sent": {
        "label": "Mailer Sent",
        "count_aliases": ["mailer sent","mailers sent","mailer","direct mail","mail sent","letters sent","mailer count"],
        "date_aliases": ["mailer date","mailer sent date","mail date","direct mail date","mailer sent on"],
    },
    "rvm": {
        "label": "RVM",
        "count_aliases": ["rvm","rvm sent","rvm count","ringless voicemail","ringless voicemail sent"],
        "date_aliases": ["rvm date","rvm sent date","ringless voicemail date"],
    },
    "ai_texting": {
        "label": "AI Texting",
        "count_aliases": ["ai texting","ai text","ai texts","ai texting sent","ai text sent","ai texting count"],
        "date_aliases": ["ai texting date","ai text date","ai texts date","ai texting sent date"],
    },
    "cold_calling": {
        "label": "Cold Calling",
        "count_aliases": ["cold calling","cold calls","cold call","cold calls made","cold calling count"],
        "date_aliases": ["cold calling date","cold call date","cold calls date"],
    },
    "neutral_postcard": {
        "label": "Neutral Postcard",
        "count_aliases": ["neutral postcard","neutral postcards","neutral postcard sent","neutral postcard count"],
        "date_aliases": ["neutral postcard date","neutral postcards date","neutral postcard sent date"],
    },
}


def _marketing_count(value, has_date: bool) -> float:
    if value is None or pd.isna(value) or str(value).strip() == "":
        return 1.0 if has_date else 0.0
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    text = str(value).strip().casefold()
    if text in {"yes","y","true","sent","done","x","si","sí","1"}:
        return 1.0
    if text in {"no","n","false","0","none","na","n/a"}:
        return 0.0
    m = re.search(r"[-+]?\d[\d,]*(?:\.\d+)?", text)
    if m:
        try:
            return max(0.0, float(m.group(0).replace(",", "")))
        except Exception:
            pass
    return 1.0 if has_date else 0.0


def _marketing_date(value) -> str:
    if value is None or pd.isna(value) or str(value).strip() == "":
        return ""
    try:
        dt = pd.to_datetime(value, errors="coerce")
        if pd.notna(dt):
            return dt.date().isoformat()
    except Exception:
        pass
    return str(value).strip()[:80]


def read_marketing_activity_file(path: Path) -> tuple[list[dict], dict]:
    df = _read_tabular_file(path)
    state_col = find_column(df, ["state","property state","prop state","estado","st"])
    county_col = find_column(df, ["county","county name","property county","condado"])
    if not state_col or not county_col:
        raise ValueError("Marketing Activity requires STATE and COUNTY columns.")

    channel_cols = {}
    for key, cfg in MARKETING_CHANNELS.items():
        channel_cols[key] = (
            find_column(df, cfg["count_aliases"]),
            find_column(df, cfg["date_aliases"]),
        )
    if not any(c or d for c, d in channel_cols.values()):
        raise ValueError("No supported marketing columns were found. Add Mailer Sent/Date, RVM/Date, AI Texting/Date, Cold Calling/Date, or Neutral Postcard/Date.")

    aggregated = {}
    valid_rows = 0
    for _, row in df.iterrows():
        state, sf = normalize_state(row.get(state_col, ""))
        county = normalize_county(row.get(county_col, ""))
        if not sf or not county:
            continue
        valid_rows += 1
        for channel, (count_col, date_col) in channel_cols.items():
            raw_date = row.get(date_col) if date_col else None
            date = _marketing_date(raw_date)
            count = _marketing_count(row.get(count_col) if count_col else None, bool(date))
            if count <= 0:
                continue
            key = (sf, county.casefold(), channel, date)
            rec = aggregated.setdefault(key, {
                "state": state, "state_fips": sf, "county": county, "county_key": county.casefold(),
                "channel": channel, "date": date, "count": 0.0,
            })
            rec["count"] += count

    records = list(aggregated.values())
    for r in records:
        if abs(r["count"] - round(r["count"])) < 1e-9:
            r["count"] = int(round(r["count"]))
        r["event_key"] = f'{r["state_fips"]}|{r["county_key"]}|{r["channel"]}|{r["date"]}'
    return records, {
        "rows_uploaded": int(len(df)), "valid_rows": valid_rows, "events": len(records),
        "states": sorted(set(r["state"] for r in records)),
        "channels": sorted(set(r["channel"] for r in records)),
    }


def merge_marketing_activity(existing: list[dict], incoming: list[dict]) -> tuple[list[dict], dict]:
    by_key = {str(r.get("event_key") or ""): dict(r) for r in existing if r.get("event_key")}
    new = updated = unchanged = 0
    for r in incoming:
        key = r["event_key"]
        old = by_key.get(key)
        if old is None:
            new += 1
        elif old == r:
            unchanged += 1
        else:
            updated += 1
        by_key[key] = dict(r)
    merged = sorted(by_key.values(), key=lambda r: (r.get("state",""), r.get("county",""), r.get("channel",""), r.get("date","")))
    return merged, {"new": new, "updated": updated, "unchanged": unchanged}


def build_marketing_activity_summary(records: list[dict]) -> list[dict]:
    grouped = {}
    for r in records:
        k = (r.get("state_fips"), r.get("county_key"))
        g = grouped.setdefault(k, {"state":r.get("state"),"state_fips":r.get("state_fips"),"county":r.get("county"),"county_key":r.get("county_key")})
        ch = r.get("channel")
        if ch not in MARKETING_CHANNELS:
            continue
        count_key = f"marketing_{ch}_count"
        dates_key = f"marketing_{ch}_dates"
        g[count_key] = float(g.get(count_key, 0) or 0) + float(r.get("count", 0) or 0)
        g.setdefault(dates_key, []).append({"date": r.get("date") or "", "count": r.get("count", 0) or 0})
    out=[]
    for g in grouped.values():
        for ch in MARKETING_CHANNELS:
            ck=f"marketing_{ch}_count"; dk=f"marketing_{ch}_dates"
            if ck in g and abs(g[ck]-round(g[ck]))<1e-9: g[ck]=int(round(g[ck]))
            if dk in g:
                g[dk]=sorted(g[dk], key=lambda x: x.get("date") or "", reverse=True)
        out.append(g)
    return out


def merge_marketing_activity_into_counties(project: dict, summary: list[dict]) -> list[dict]:
    counties = project.setdefault("counties", [])
    fields=[]
    for ch in MARKETING_CHANNELS:
        fields += [f"marketing_{ch}_count", f"marketing_{ch}_dates"]
    for c in counties:
        for field in fields:
            c.pop(field, None)
    by_id={(c.get("state_fips"),c.get("county_key")):c for c in counties}
    for m in summary:
        key=(m.get("state_fips"),m.get("county_key"))
        c=by_id.get(key)
        if c is None:
            c={"county":m.get("county"),"county_key":m.get("county_key"),"state":m.get("state"),"state_fips":m.get("state_fips"),"status":"Marketing data","date":"","notes":"","priority":"","assigned_to":"","next_review":"","str":"","str_value":None}
            for band in STR_COLUMN_ALIASES:
                c[band]=""; c[f"{band}_value"]=None
            counties.append(c); by_id[key]=c
        for field in fields:
            if field in m: c[field]=m[field]
    counties.sort(key=lambda c:(c.get("state",""),c.get("county","")))
    return counties


@app.post("/api/projects/<project_id>/marketing-activity")
def upload_marketing_activity(project_id: str):
    try:
        project=load_project(project_id)
    except (FileNotFoundError,ValueError):
        return jsonify({"error":"Map not found"}),404
    file=request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error":"Select a marketing activity file."}),400
    if Path(file.filename).suffix.lower() not in {".xlsx",".xlsm",".csv"}:
        return jsonify({"error":"Use an .xlsx, .xlsm, or .csv file."}),400
    dest=UPLOAD_DIR / f"{project_id}_marketing_{secure_filename(file.filename)}"
    file.save(dest)
    try:
        incoming,meta=read_marketing_activity_file(dest)
        existing=storage.load_marketing_activity(project_id)
        merged,merge_meta=merge_marketing_activity(existing,incoming)
        storage.save_marketing_activity(project_id,merged)
        summary=build_marketing_activity_summary(merged)
        counties=merge_marketing_activity_into_counties(project,summary)
    except Exception as exc:
        return jsonify({"error":str(exc)}),400
    analytics=project.setdefault("marketing_activity",{})
    files=analytics.setdefault("source_files",[])
    files.append({"name":file.filename,"uploaded_at":now_iso(),"rows":meta.get("rows_uploaded"),"events":meta.get("events")})
    analytics.update({
        "loaded":True,"states":sorted(set(r.get("state") for r in merged if r.get("state"))),
        "events":len(merged),"last_upload_at":now_iso(),"last_upload_rows":meta.get("rows_uploaded"),
        "last_upload_valid":meta.get("valid_rows"),"channels":meta.get("channels",[]),
    })
    project["counties"]=counties
    save_project(project)
    socketio.emit("counties_updated", {"counties":counties}, to=project_id)
    return jsonify({"ok":True,"counties":counties,"analytics":analytics,"upload":{**meta,**merge_meta}})


def _market_state_county(row, df):
    state_col = find_column(df, ["state", "st", "property state", "address state", "region"])
    county_col = find_column(df, ["county", "county name", "property county", "condado"])
    state = str(row.get(state_col, "") if state_col else "").strip()
    county = str(row.get(county_col, "") if county_col else "").strip()

    # Land.com-style coded downloads: structured JSON and address text.
    raw_json = row.get("bc7c8e2") if "bc7c8e2" in df.columns else None
    if (not state or not county) and raw_json is not None and not pd.isna(raw_json):
        try:
            data = json.loads(str(raw_json))
            addr = data.get("address") or {}
            state = state or str(addr.get("addressRegion") or "").strip()
        except Exception:
            pass
    addr_text = ""
    if "_28d22f4" in df.columns:
        v = row.get("_28d22f4")
        if v is not None and not pd.isna(v):
            addr_text = str(v)
    if not county and addr_text:
        m = re.search(r"([^,]+?)\s+County\s*$", addr_text, flags=re.I)
        if m:
            county = m.group(1).strip()
    if not state and addr_text:
        # find 2-letter state token followed by ZIP/comma
        m = re.search(r",\s*([A-Z]{2})\s*,?\s*\d{5}(?:-\d{4})?", addr_text)
        if m:
            state = m.group(1)

    state, sf = normalize_state(state)
    county = normalize_county(county)
    return state, sf, county


def read_market_file(path: Path, status: str) -> tuple[list[dict], dict]:
    df = _read_tabular_file(path)
    price_col = find_column(df, ["price", "list price", "sold price", "sale price", "asking price", "listing price"])
    acres_col = find_column(df, ["acres", "acreage", "lot acres", "size acres", "land acres"])
    url_col = find_column(df, ["url", "listing url", "property url", "link", "href"])
    address_col = find_column(df, ["address", "property address", "listing address", "location"])
    # Known Land.com encoded columns.
    if price_col is None and "_47a280d" in df.columns: price_col = "_47a280d"
    if acres_col is None and "_28423b5" in df.columns: acres_col = "_28423b5"
    if url_col is None and "_1cd8ad9 href" in df.columns: url_col = "_1cd8ad9 href"
    if address_col is None and "_28d22f4" in df.columns: address_col = "_28d22f4"

    records = {}
    skipped = 0
    for _, row in df.iterrows():
        state, sf, county = _market_state_county(row, df)
        acres = _acres_float(row.get(acres_col)) if acres_col else None
        price = _money_float(row.get(price_col)) if price_col else None
        if not sf or not county:
            skipped += 1
            continue
        url = ""
        if url_col:
            v = row.get(url_col)
            url = "" if v is None or pd.isna(v) else str(v).strip()
        address = ""
        if address_col:
            v = row.get(address_col)
            address = "" if v is None or pd.isna(v) else str(v).strip()
        # Prefer stable listing URL. Fallback is normalized address + acres + state/county.
        if url:
            key = f"url|{url.casefold()}"
        else:
            key = "|".join([state, county.casefold(), _norm_owner_piece(address), str(acres or "")])
        records[key] = {
            "market_key": key,
            "status": status,
            "state": state,
            "state_fips": sf,
            "county": county,
            "county_key": county.casefold(),
            "acres": acres,
            "price": price,
            "address": address,
            "source_file": path.name,
            "updated_at": now_iso(),
        }
    out = list(records.values())
    if not out:
        raise ValueError("No valid market rows with state/county could be read from this file.")
    return out, {"rows_uploaded": int(len(df)), "valid_rows": len(out), "skipped_rows": skipped}


def merge_market_records(existing: list[dict], incoming: list[dict], status: str) -> tuple[list[dict], dict]:
    # ACTIVE is a snapshot by state: replace prior active inventory for states in the upload.
    # SOLD is historical/event data: upsert by listing key and keep prior sold records.
    incoming_states = {r.get("state") for r in incoming if r.get("state")}
    if status == "active":
        kept = [r for r in existing if not (r.get("status") == "active" and r.get("state") in incoming_states)]
        merged = kept + incoming
        return merged, {"replaced_active_states": sorted(incoming_states), "added": len(incoming)}
    by_key = {str(r.get("market_key") or ""): dict(r) for r in existing if r.get("market_key")}
    added = updated = 0
    for r in incoming:
        key = r["market_key"]
        if key in by_key: updated += 1
        else: added += 1
        by_key[key] = r
    return list(by_key.values()), {"added": added, "updated": updated}


def calculate_market_metrics(records: list[dict], min_acres=None, max_acres=None) -> list[dict]:
    def keep(r):
        a = r.get("acres")
        if min_acres is not None and (a is None or a < min_acres): return False
        if max_acres is not None and (a is None or a > max_acres): return False
        return True
    grouped = {}
    for r in records:
        if not keep(r): continue
        k = (r.get("state_fips"), r.get("county_key"))
        g = grouped.setdefault(k, {"state":r.get("state"),"state_fips":r.get("state_fips"),"county":r.get("county"),"county_key":r.get("county_key"),"active":[],"sold":[]})
        g[r.get("status") if r.get("status") in {"active","sold"} else "active"].append(r)
    out=[]
    for g in grouped.values():
        active=g["active"]; sold=g["sold"]
        ap=[r.get("price") for r in active if isinstance(r.get("price"),(int,float))]
        sp=[r.get("price") for r in sold if isinstance(r.get("price"),(int,float))]
        ac=len(active); sc=len(sold)
        out.append({
            "state":g["state"],"state_fips":g["state_fips"],"county":g["county"],"county_key":g["county_key"],
            "market_active_count":ac,"market_sold_count":sc,
            "market_str": round((sc/ac)*100,2) if ac else None,
            "market_avg_active_price": round(sum(ap)/len(ap),2) if ap else None,
            "market_avg_sold_price": round(sum(sp)/len(sp),2) if sp else None,
        })
    return out


@app.post("/api/projects/<project_id>/market/<status>")
def upload_market_data(project_id: str, status: str):
    status = str(status or "").lower().strip()
    if status not in {"active","sold"}:
        return jsonify({"error":"Status must be active or sold."}), 400
    try:
        project = load_project(project_id)
    except (FileNotFoundError, ValueError):
        return jsonify({"error":"Map not found"}), 404
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error":"Select a CSV/Excel market file."}), 400
    ext=Path(file.filename).suffix.lower()
    if ext not in {".csv",".xlsx",".xlsm"}:
        return jsonify({"error":"Use .csv, .xlsx, or .xlsm."}),400
    dest=UPLOAD_DIR / f"{project_id}_market_{status}_{secure_filename(file.filename)}"
    file.save(dest)
    try:
        incoming, meta = read_market_file(dest, status)
        existing = storage.load_market_records(project_id)
        merged, merge_meta = merge_market_records(existing, incoming, status)
        storage.save_market_records(project_id, merged)
    except Exception as exc:
        return jsonify({"error":str(exc)}),400
    analytics=project.setdefault("market_analytics", {})
    key="active_files" if status=="active" else "sold_files"
    arr=list(analytics.get(key,[]))
    if file.filename not in arr: arr.append(file.filename)
    analytics[key]=arr[-50:]
    analytics.update({
        "loaded":True,
        "active_count":sum(1 for r in merged if r.get("status")=="active"),
        "sold_count":sum(1 for r in merged if r.get("status")=="sold"),
        "states":sorted({r.get("state") for r in merged if r.get("state")}),
        "last_upload_status":status,
        "last_upload_at":now_iso(),
        "last_upload_rows":meta.get("rows_uploaded"),
        "last_upload_valid":meta.get("valid_rows"),
    })
    save_project(project)
    return jsonify({"ok":True,"analytics":analytics,"upload":{**meta,**merge_meta}})


@app.get("/api/projects/<project_id>/market-metrics")
def market_metrics(project_id: str):
    try:
        load_project(project_id)
    except (FileNotFoundError, ValueError):
        return jsonify({"error":"Map not found"}),404
    try:
        min_acres = request.args.get("min_acres")
        max_acres = request.args.get("max_acres")
        min_acres = float(min_acres) if min_acres not in {None,""} else None
        max_acres = float(max_acres) if max_acres not in {None,""} else None
    except ValueError:
        return jsonify({"error":"Invalid acreage range"}),400
    records=storage.load_market_records(project_id)
    metrics=calculate_market_metrics(records,min_acres,max_acres)
    return jsonify({"metrics":metrics,"counties":len(metrics),"min_acres":min_acres,"max_acres":max_acres})


@app.post("/api/projects/<project_id>/properties")
def upload_properties(project_id: str):
    try:
        project = load_project(project_id)
    except (FileNotFoundError, ValueError):
        return jsonify({"error": "Map not found"}), 404
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "Select a property CSV/Excel file."}), 400
    ext = Path(file.filename).suffix.lower()
    if ext not in {".csv", ".xlsx", ".xlsm"}:
        return jsonify({"error": "Use .csv, .xlsx, or .xlsm."}), 400
    dest = UPLOAD_DIR / f"{project_id}_properties_{secure_filename(file.filename)}"
    file.save(dest)
    try:
        incoming, upload_meta = read_property_file(dest)
        existing = storage.load_property_records(project_id)
        master_records, upsert = merge_property_records(existing, incoming)
        metrics, points, master_meta = build_property_metrics(master_records)
        storage.save_property_records(project_id, master_records)
        storage.save_property_points(project_id, points)
        counties = merge_property_metrics_into_counties(project, metrics)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400

    analytics = project.setdefault("property_analytics", {})
    source_files = list(analytics.get("source_files", []))
    if file.filename not in source_files:
        source_files.append(file.filename)
    analytics.update({
        "loaded": True,
        "source_files": source_files[-50:],
        "states": master_meta["states"],
        "total_properties": master_meta["total_properties"],
        "unique_owners": master_meta["unique_owners"],
        "portfolio_owners": master_meta["portfolio_owners"],
        "county_count": master_meta["county_count"],
        "point_count": master_meta["point_count"],
        "last_upload_rows": upload_meta["rows_uploaded"],
        "last_upload_valid_refs": upload_meta["valid_refs"],
        "last_upload_new": upsert["new"],
        "last_upload_updated": upsert["updated"],
        "last_upload_unchanged": upsert["unchanged"],
        "last_upload_duplicate_refs": upload_meta["duplicate_refs_in_file"],
        "last_upload_missing_ref": upload_meta["skipped_no_ref"],
        "last_upload_at": now_iso(),
        "property_key": "REF",
        "owner_key": "normalized owner name + mailing address",
    })
    project["counties"] = counties
    save_project(project)
    socketio.emit("counties_updated", {"counties": counties, "updated_at": project["updated_at"]}, to=project_id)
    return jsonify({"ok": True, "counties": counties, "analytics": analytics, "metrics_count": len(metrics), "upload": {**upload_meta, **upsert}})


@app.get("/api/projects/<project_id>/property-points")
def property_points(project_id: str):
    try: load_project(project_id)
    except (FileNotFoundError,ValueError): return jsonify({"error":"Map not found"}),404
    state=(request.args.get("state") or "").upper().strip()
    try:
        west=float(request.args.get("west","-180")); south=float(request.args.get("south","-90")); east=float(request.args.get("east","180")); north=float(request.args.get("north","90"))
    except ValueError: return jsonify({"error":"Invalid bbox"}),400
    points=storage.load_property_points(project_id)
    selected=[]
    for p in points:
        if state and p.get("state")!=state: continue
        lat=p.get("lat"); lon=p.get("lon")
        if lat is None or lon is None or not (south<=lat<=north and west<=lon<=east): continue
        selected.append(p)
        if len(selected)>=25000: break
    return jsonify({"points":selected,"count":len(selected),"truncated":len(selected)>=25000})


@app.get("/")
def index():
    storage_error = ""
    try:
        raw_projects = storage.list_projects()
    except RuntimeError as exc:
        storage_error = str(exc)
        raw_projects = []
    projects = [
        {"id": p["id"], "name": p.get("name", "Map"), "updated_at": p.get("updated_at", "")}
        for p in raw_projects
    ]
    return render_template(
        "index.html",
        projects=projects,
        storage_mode=storage.mode,
        storage_error=storage_error,
    )


@app.get("/health")
def health():
    return jsonify({"ok": True, "storage": storage.mode})


@app.post("/projects")
def create_project():
    name = (request.form.get("name") or "County Map").strip()
    project_id = uuid.uuid4().hex[:12]
    project = {
        "id": project_id,
        "name": name,
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "schema_version": PROJECT_SCHEMA_VERSION,
        "counties": [],
        "drawings": {"type": "FeatureCollection", "features": []},
        "view_settings": {
            "state_filter": "", "str_metric": "str_value", "search_filter": "", "color_metric": "str_value",
            "property_color_mode": "automatic", "property_thresholds": [25,100,250,500],
            "property_point_style": {"size": 4, "color": "#22c55e", "opacity": 0.75, "outline_color": "#0f172a", "outline_width": 1},
            "layers": {"counties": True, "county_labels": True, "str_colors": True, "drawings": True, "drawing_labels": True, "property_points": True},
            "layer_order": ["counties", "drawings", "drawing_labels", "county_labels", "property_points"],
            "county_card_fields": ["average_str","property_count","unique_owner_count","portfolio_owner_count","portfolio_property_count","total_acreage","average_acreage","market_active_count","market_sold_count","market_str","market_avg_active_price","market_avg_sold_price","str_by_acreage","mailer_sent","rvm","ai_texting","cold_calling","neutral_postcard"],
        },
    }
    save_project(project)
    return redirect(url_for("map_view", project_id=project_id))


@app.get("/map/<project_id>")
def map_view(project_id: str):
    try:
        project = load_project(project_id)
    except (FileNotFoundError, ValueError):
        return "Map not found", 404
    return render_template("map.html", project=project, storage_mode=storage.mode)


@app.get("/api/projects/<project_id>")
def project_data(project_id: str):
    try:
        return jsonify(load_project(project_id))
    except (FileNotFoundError, ValueError):
        return jsonify({"error": "not found"}), 404


@app.post("/api/projects/<project_id>/excel")
def upload_excel(project_id: str):
    try:
        project = load_project(project_id)
    except (FileNotFoundError, ValueError):
        return jsonify({"error": "Map not found"}), 404
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "Select an Excel file."}), 400
    ext = Path(file.filename).suffix.lower()
    if ext not in {".xlsx", ".xlsm", ".csv"}:
        return jsonify({"error": "Use an .xlsx, .xlsm, or .csv file."}), 400
    filename = f"{project_id}_{secure_filename(file.filename)}"
    dest = UPLOAD_DIR / filename
    file.save(dest)
    try:
        imported = read_counties_csv(dest) if ext == ".csv" else read_counties_excel(dest)
        counties = merge_existing_notes(imported, project.get("counties", []))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400
    project["counties"] = counties
    project["source_file"] = file.filename
    save_project(project)
    socketio.emit("counties_updated", {"counties": counties, "updated_at": project["updated_at"]}, to=project_id)
    return jsonify({"ok": True, "count": len(counties), "counties": counties})


@app.post("/api/projects/<project_id>/counties/activate")
def activate_county(project_id: str):
    try:
        project = load_project(project_id)
    except (FileNotFoundError, ValueError):
        return jsonify({"error": "Map not found"}), 404
    data = request.get_json(silent=True) or {}
    state_fips = str(data.get("state_fips") or "").zfill(2)
    county = normalize_county(data.get("county"))
    state = FIPS_STATES.get(state_fips, "")
    if not state or not county:
        return jsonify({"error": "Invalid county"}), 400
    county_key = county.casefold()
    for existing in project.get("counties", []):
        if existing.get("state_fips") == state_fips and existing.get("county_key") == county_key:
            return jsonify({"ok": True, "county": existing, "already_active": True})
    county_data = {
        "county": county, "county_key": county_key, "state": state, "state_fips": state_fips,
        "status": "Manually selected", "date": "", "notes": "", "priority": "",
        "assigned_to": "", "next_review": "", "str": "", "str_value": None,
    }
    for key in STR_COLUMN_ALIASES:
        county_data[key] = ""
        county_data[f"{key}_value"] = None
    project.setdefault("counties", []).append(county_data)
    project["counties"].sort(key=lambda c: (c.get("state", ""), c.get("county", "")))
    save_project(project)
    socketio.emit("counties_updated", {"counties": project["counties"], "updated_at": project["updated_at"]}, to=project_id)
    return jsonify({"ok": True, "county": county_data, "counties": project["counties"]})


@app.post("/api/projects/<project_id>/counties/notes")
def save_county_note(project_id: str):
    try:
        project = load_project(project_id)
    except (FileNotFoundError, ValueError):
        return jsonify({"error": "Map not found"}), 404

    data = request.get_json(silent=True) or {}
    state_fips = str(data.get("state_fips", "")).strip()
    county_key = str(data.get("county_key", "")).strip().casefold()
    notes = str(data.get("notes", "")).strip()[:5000]
    priority = str(data.get("priority", "")).strip()[:50]
    assigned_to = str(data.get("assigned_to", "")).strip()[:200]
    next_review = str(data.get("next_review", "")).strip()[:100]
    if not state_fips or not county_key:
        return jsonify({"error": "Invalid county"}), 400

    updated = None
    for county in project.get("counties", []):
        if str(county.get("state_fips", "")) == state_fips and str(county.get("county_key", "")).casefold() == county_key:
            county["notes"] = notes
            county["priority"] = priority
            county["assigned_to"] = assigned_to
            county["next_review"] = next_review
            updated = county
            break
    if updated is None:
        return jsonify({"error": "County was not found in the Excel file"}), 404

    save_project(project)
    socketio.emit(
        "county_note_updated",
        {"county": updated, "sender": data.get("sender"), "updated_at": project["updated_at"]},
        to=project_id,
    )
    return jsonify({"ok": True, "county": updated, "updated_at": project["updated_at"]})


@app.post("/api/projects/<project_id>/drawings")
def save_drawings(project_id: str):
    try:
        project = load_project(project_id)
    except (FileNotFoundError, ValueError):
        return jsonify({"error": "Map not found"}), 404
    data = request.get_json(silent=True) or {}
    drawings = data.get("drawings")
    if not isinstance(drawings, dict) or drawings.get("type") != "FeatureCollection":
        return jsonify({"error": "Invalid GeoJSON"}), 400
    project["drawings"] = drawings
    save_project(project)
    socketio.emit("drawings_updated", {"drawings": drawings, "sender": data.get("sender")}, to=project_id)
    return jsonify({"ok": True, "updated_at": project["updated_at"]})


@app.post("/api/projects/<project_id>/settings")
def save_project_settings(project_id: str):
    try:
        project = load_project(project_id)
    except (FileNotFoundError, ValueError):
        return jsonify({"error": "Map not found"}), 404
    data = request.get_json(silent=True) or {}
    settings = data.get("view_settings")
    if not isinstance(settings, dict):
        return jsonify({"error": "Invalid settings"}), 400
    allowed_layers = {"counties", "county_labels", "str_colors", "drawings", "drawing_labels", "property_points"}
    allowed_card_fields = {"average_str","property_count","unique_owner_count","portfolio_owner_count","portfolio_property_count","total_acreage","average_acreage","market_active_count","market_sold_count","market_str","market_avg_active_price","market_avg_sold_price","str_by_acreage","mailer_sent","rvm","ai_texting","cold_calling","neutral_postcard","status"}
    clean = {
        "state_filter": str(settings.get("state_filter") or "")[:2].upper(),
        "str_min": settings.get("str_min") if isinstance(settings.get("str_min"), (int, float)) else None,
        "str_max": settings.get("str_max") if isinstance(settings.get("str_max"), (int, float)) else None,
        "search_filter": str(settings.get("search_filter") or "")[:200],
        "str_metric": str(settings.get("str_metric") or "str_value")[:40],
        "color_metric": str(settings.get("color_metric") or settings.get("str_metric") or "str_value")[:50],
        "property_color_mode": "custom" if settings.get("property_color_mode") == "custom" else "automatic",
        "property_thresholds": [float(x) for x in (settings.get("property_thresholds") or [25,100,250,500])[:4] if isinstance(x,(int,float))],
        "property_point_style": {
            "size": max(1.0, min(20.0, float((settings.get("property_point_style") or {}).get("size", 4) or 4))),
            "color": str((settings.get("property_point_style") or {}).get("color") or "#22c55e")[:20],
            "opacity": max(0.05, min(1.0, float((settings.get("property_point_style") or {}).get("opacity", 0.75) or 0.75))),
            "outline_color": str((settings.get("property_point_style") or {}).get("outline_color") or "#0f172a")[:20],
            "outline_width": max(0.0, min(5.0, float((settings.get("property_point_style") or {}).get("outline_width", 1) or 1))),
        },
        "layers": {k: bool((settings.get("layers") or {}).get(k, True)) for k in allowed_layers},
        "layer_order": [k for k in (settings.get("layer_order") or ["counties","drawings","drawing_labels","county_labels","property_points"]) if k in {"counties","drawings","drawing_labels","county_labels","property_points"}],
        "county_card_fields": [k for k in (settings.get("county_card_fields") or []) if k in allowed_card_fields],
    }
    if not clean["county_card_fields"]:
        clean["county_card_fields"] = ["average_str","property_count","portfolio_owner_count","mailer_sent"]
    # Ensure every reorderable layer exists exactly once.
    for _k in ["counties","drawings","drawing_labels","county_labels","property_points"]:
        if _k not in clean["layer_order"]:
            clean["layer_order"].append(_k)
    project["view_settings"] = clean
    save_project(project)
    socketio.emit("settings_updated", {"view_settings": clean, "sender": data.get("sender")}, to=project_id)
    return jsonify({"ok": True, "view_settings": clean, "updated_at": project["updated_at"]})


@app.post("/api/projects/<project_id>/rename")
def rename_project(project_id: str):
    try:
        project = load_project(project_id)
    except (FileNotFoundError, ValueError):
        return jsonify({"error": "Map not found"}), 404
    data = request.get_json(silent=True) or {}
    name = str(data.get("name") or "").strip()[:150]
    if not name:
        return jsonify({"error": "Enter a project name"}), 400
    project["name"] = name
    save_project(project)
    socketio.emit("project_renamed", {"name": name}, to=project_id)
    return jsonify({"ok": True, "name": name})


@app.post("/api/projects/<project_id>/duplicate")
def duplicate_project(project_id: str):
    try:
        source = load_project(project_id)
    except (FileNotFoundError, ValueError):
        return jsonify({"error": "Map not found"}), 404
    clone = json.loads(json.dumps(source))
    clone["id"] = uuid.uuid4().hex[:12]
    clone["name"] = f"{source.get('name', 'Map')} Copy"
    clone["created_at"] = now_iso()
    clone["updated_at"] = now_iso()
    save_project(clone)
    try:
        storage.save_property_records(clone["id"], storage.load_property_records(project_id))
        storage.save_property_points(clone["id"], storage.load_property_points(project_id))
        storage.save_market_records(clone["id"], storage.load_market_records(project_id))
        storage.save_marketing_activity(clone["id"], storage.load_marketing_activity(project_id))
    except Exception:
        # Project duplication should still succeed even if an old map has no property store.
        pass
    return jsonify({"ok": True, "id": clone["id"], "url": url_for("map_view", project_id=clone["id"])})


@app.delete("/api/projects/<project_id>")
def delete_project(project_id: str):
    try:
        load_project(project_id)
        storage.delete(project_id)
    except FileNotFoundError:
        return jsonify({"error": "Map not found"}), 404
    except ValueError:
        return jsonify({"error": "Invalid project id"}), 400
    return jsonify({"ok": True})


@socketio.on("join_project")
def join_project_event(data):
    project_id = str((data or {}).get("project_id", ""))
    try:
        load_project(project_id)
    except Exception:
        return
    join_room(project_id)
    emit("joined", {"project_id": project_id})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    production = os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("RENDER")
    if not production:
        threading.Timer(1.2, lambda: webbrowser.open(f"http://127.0.0.1:{port}")).start()
    socketio.run(
        app,
        host="0.0.0.0",
        port=port,
        debug=not bool(production),
        use_reloader=False,
        allow_unsafe_werkzeug=True,
    )
